"""Excel exporter for Facebook Comments."""

from __future__ import annotations

import io
import logging

from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

from .models import CommentData, EXCEL_HEADERS

log = logging.getLogger(__name__)

HEADER_FONT = Font(bold=True, color="FFFFFF", size=11)
HEADER_FILL = PatternFill(start_color="1877F2", end_color="1877F2", fill_type="solid")
HEADER_ALIGNMENT = Alignment(horizontal="center", vertical="center", wrap_text=True)
HEADER_BORDER = Border(
    bottom=Side(style="thin", color="000000"),
    right=Side(style="thin", color="CCCCCC"),
)
DATA_ALIGNMENT = Alignment(vertical="top", wrap_text=True)

class ExcelExporter:
    """Exports comments to an Excel workbook."""

    def __init__(self):
        self._workbook = Workbook()
        self._sheet = self._workbook.active
        self._sheet.title = "Comments"
        self._current_row = 1
        self._total_rows = 0
        self._init_sheet()

    def _init_sheet(self) -> None:
        for col_idx, header in enumerate(EXCEL_HEADERS, 1):
            cell = self._sheet.cell(row=1, column=col_idx, value=header)
            cell.font = HEADER_FONT
            cell.fill = HEADER_FILL
            cell.alignment = HEADER_ALIGNMENT
            cell.border = HEADER_BORDER

        column_widths = {
            1: 40,  # Post URL
            2: 20,  # Post Author Name
            3: 60,  # Post Text
            4: 10,  # Comment #
            5: 12,  # Type
            6: 18,  # Comment ID
            7: 18,  # Post ID
            8: 20,  # Author Name
            9: 40,  # Author URL
            10: 60, # Comment Text
            11: 12, # Reaction Count
            12: 12, # Reply Count
            13: 10, # Is Reply
            14: 18, # Parent Comment ID
            15: 20, # Scraped At
        }
        for col, width in column_widths.items():
            self._sheet.column_dimensions[get_column_letter(col)].width = width

        self._sheet.freeze_panes = "A2"

    def add_comment(self, comment: CommentData) -> None:
        self._current_row += 1
        self._total_rows += 1
        
        row_data = comment.to_excel_row()
        for col_idx, value in enumerate(row_data, 1):
            cell = self._sheet.cell(row=self._current_row, column=col_idx, value=value)
            cell.alignment = DATA_ALIGNMENT
            
            # Make Post URL (col 1) and Author URL (col 9) clickable
            if col_idx in (1, 9) and value: 
                try:
                    cell.hyperlink = value
                    cell.font = Font(color="1155CC", underline="single")
                except:
                    pass
        
        # Subtle gray background for reply rows to visually group them
        if comment.is_reply:
            reply_fill = PatternFill(start_color="F0F0F0", end_color="F0F0F0", fill_type="solid")
            for col_idx in range(1, len(row_data) + 1):
                self._sheet.cell(row=self._current_row, column=col_idx).fill = reply_fill

    def save_to_bytes(self) -> bytes:
        buffer = io.BytesIO()
        self._workbook.save(buffer)
        buffer.seek(0)
        return buffer.read()

    def get_stats(self) -> dict:
        return {
            "total_rows": self._total_rows
        }
